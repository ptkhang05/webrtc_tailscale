import asyncio
import os
import ssl
import time
from unittest.mock import patch

from openpyxl import load_workbook

from web_intercom.certs import ensure_self_signed_cert
from web_intercom.metrics import derive_client_qoe_metrics, estimate_e_model
from web_intercom.server import (
    Client,
    MAX_AUTH_FAILURES_PER_IP,
    MAX_CLIENTS,
    MAX_CONNECTIONS_PER_IP,
    MAX_OPEN_CONNECTIONS,
    RELAY_QUEUE_MAX_PACKETS,
    WebIntercomServer,
    audio_payload_size,
    audio_stream_id,
    build_ssl_context,
    build_arg_parser,
    create_app,
    load_room_key,
)


def test_parser_defaults():
    args = build_arg_parser().parse_args([])

    assert args.host == "0.0.0.0"
    assert args.port == 8443
    assert args.stats_interval == 5.0
    assert args.metrics_xlsx is None
    assert args.max_clients == MAX_CLIENTS
    assert args.max_open_connections == MAX_OPEN_CONNECTIONS
    assert args.max_connections_per_ip == MAX_CONNECTIONS_PER_IP
    assert args.max_auth_failures_per_ip == MAX_AUTH_FAILURES_PER_IP
    assert args.auth_failure_window == 60.0


def test_load_room_key_from_value():
    assert load_room_key("secret") == "secret"


def test_room_recipients_excludes_sender():
    relay = WebIntercomServer("secret")
    sender = object()
    receiver = object()
    relay.clients[sender] = type("Client", (), {"websocket": sender, "name": "a", "room": "main"})()
    relay.clients[receiver] = type("Client", (), {"websocket": receiver, "name": "b", "room": "main"})()

    recipients = asyncio.run(relay.room_recipients("main", exclude=sender))

    assert recipients == [receiver]


def test_create_app_routes(tmp_path):
    (tmp_path / "index.html").write_text("<html></html>", encoding="utf-8")
    app = create_app("secret", tmp_path, 0)
    paths = {route.resource.canonical for route in app.router.routes()}

    assert "/" in paths
    assert "/ws" in paths


def test_audio_packet_header_validation():
    assert audio_payload_size(b"SWI1" + b"\x00" * 16 + b"pcm") == 3
    assert audio_payload_size(b"BAD!" + b"\x00" * 16 + b"pcm") is None
    assert audio_payload_size(b"SWI1") is None
    packet = b"SWI1" + (123456).to_bytes(4, "big") + b"\x00" * 12 + b"pcm"
    assert audio_stream_id(packet) == 123456


def test_generated_private_key_uses_restrictive_open_mode(tmp_path):
    real_open = os.open
    modes = []

    def tracking_open(path, flags, mode=0o777):
        modes.append(mode)
        return real_open(path, flags, mode)

    with patch("web_intercom.certs.os.open", tracking_open):
        cert_path, key_path = ensure_self_signed_cert(tmp_path / "certs", ["127.0.0.1"])

    assert cert_path.exists()
    assert key_path.exists()
    assert 0o600 in modes


def test_ssl_context_requires_tls_1_2_or_newer(tmp_path):
    cert_path, key_path = ensure_self_signed_cert(tmp_path / "certs", ["127.0.0.1"])

    context = build_ssl_context(cert_path, key_path)

    assert context.minimum_version == ssl.TLSVersion.TLSv1_2


def test_connection_reservation_limits_total_and_per_ip():
    async def run() -> None:
        relay = WebIntercomServer("secret", max_open_connections=2, max_connections_per_ip=1)

        assert await relay.reserve_connection("10.0.0.2")
        assert not await relay.reserve_connection("10.0.0.2")
        assert await relay.reserve_connection("10.0.0.3")
        assert not await relay.reserve_connection("10.0.0.4")
        await relay.release_connection("10.0.0.2")
        assert await relay.reserve_connection("10.0.0.4")

    asyncio.run(run())


def test_auth_failure_rate_limit_and_clear():
    async def run() -> None:
        relay = WebIntercomServer(
            "secret",
            max_auth_failures_per_ip=2,
            auth_failure_window_seconds=60,
        )
        remote_ip = "10.0.0.2"

        assert not await relay.is_auth_rate_limited(remote_ip)
        await relay.record_auth_failure(remote_ip)
        await relay.record_auth_failure(remote_ip)
        assert await relay.is_auth_rate_limited(remote_ip)
        await relay.clear_auth_failures(remote_ip)
        assert not await relay.is_auth_rate_limited(remote_ip)

    asyncio.run(run())


def test_stats_loop_logs_and_continues_after_unexpected_error(capfd):
    async def run() -> int:
        relay = WebIntercomServer("secret")
        original_sample = relay.metrics.sample
        calls = 0

        def flaky_sample(active_clients):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise ValueError("boom")
            return original_sample(active_clients)

        relay.metrics.sample = flaky_sample
        task = asyncio.create_task(relay.stats_loop(0.01, None))
        await asyncio.sleep(0.05)
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass
        return calls

    calls = asyncio.run(run())
    output = capfd.readouterr().out

    assert calls >= 2
    assert "[stats_loop] unexpected error: ValueError: boom" in output


def test_handle_audio_does_not_block_on_slow_recipient():
    class SlowWebSocket:
        async def send_bytes(self, data: bytes) -> None:
            await asyncio.sleep(1)

    async def run() -> None:
        relay = WebIntercomServer("secret")
        sender_ws = object()
        slow_ws = SlowWebSocket()
        sender = Client(sender_ws, "client-0001", "alice", "main", time.monotonic())
        slow = Client(slow_ws, "client-0002", "bob", "main", time.monotonic())
        relay.clients[sender_ws] = sender
        relay.clients[slow_ws] = slow
        packet = b"SWI1" + b"\x00" * 16 + b"pcm"
        relay.start_relay_worker(slow)

        try:
            started_at = time.perf_counter()
            await relay.handle_audio(sender, packet)
            elapsed = time.perf_counter() - started_at
            await asyncio.sleep(0.05)
            sample = relay.metrics.sample([])
        finally:
            await relay.stop_relay_worker(slow)

        assert elapsed < 0.05
        assert sample["rx_audio_packets"] == 1
        assert sample["relay_send_drops"] == 1

    asyncio.run(run())


def test_handle_audio_registers_sender_stream_id():
    async def run() -> None:
        relay = WebIntercomServer("secret")
        sender_ws = object()
        sender = Client(sender_ws, "client-0001", "alice", "main", time.monotonic())
        relay.clients[sender_ws] = sender
        packet = b"SWI1" + (42).to_bytes(4, "big") + b"\x00" * 12 + b"pcm"

        await relay.handle_audio(sender, packet)

        assert sender.stream_ids == {42}

    asyncio.run(run())


def test_handle_audio_drops_when_recipient_relay_queue_is_full():
    class HealthyWebSocket:
        async def send_bytes(self, data: bytes) -> None:
            raise AssertionError("send should not be scheduled when the relay queue is full")

    async def run() -> None:
        relay = WebIntercomServer("secret")
        sender_ws = object()
        recipient_ws = HealthyWebSocket()
        sender = Client(sender_ws, "client-0001", "alice", "main", time.monotonic())
        sender.stream_ids.add(0)
        recipient = Client(recipient_ws, "client-0002", "bob", "main", time.monotonic())
        for _ in range(RELAY_QUEUE_MAX_PACKETS):
            recipient.relay_queue.put_nowait((b"old", 3))
        relay.clients[sender_ws] = sender
        relay.clients[recipient_ws] = recipient
        packet = b"SWI1" + b"\x00" * 16 + b"pcm"

        await relay.handle_audio(sender, packet)
        sample = relay.metrics.sample([])

        assert len(relay.relay_tasks) == 0
        assert recipient.relay_queue.qsize() == RELAY_QUEUE_MAX_PACKETS
        assert sample["rx_audio_packets"] == 1
        assert sample["relay_send_drops"] == 1

    asyncio.run(run())


def test_relay_worker_survives_unexpected_send_exception():
    class FlakyWebSocket:
        def __init__(self) -> None:
            self.calls = 0
            self.sent = []

        async def send_bytes(self, data: bytes) -> None:
            self.calls += 1
            if self.calls == 1:
                raise ValueError("unexpected transport state")
            self.sent.append(data)

    async def run() -> None:
        relay = WebIntercomServer("secret")
        websocket = FlakyWebSocket()
        recipient = Client(websocket, "client-0002", "bob", "main", time.monotonic())
        relay.start_relay_worker(recipient)
        first = b"SWI1" + b"\x00" * 16 + b"one"
        second = b"SWI1" + b"\x00" * 16 + b"two"

        try:
            recipient.relay_queue.put_nowait((first, 3))
            await asyncio.wait_for(recipient.relay_queue.join(), timeout=1)
            assert recipient.relay_worker is not None
            assert not recipient.relay_worker.done()
            recipient.relay_queue.put_nowait((second, 3))
            await asyncio.wait_for(recipient.relay_queue.join(), timeout=1)
            sample = relay.metrics.sample([])
        finally:
            await relay.stop_relay_worker(recipient)

        assert websocket.sent == [second]
        assert sample["relay_send_drops"] == 1
        assert sample["relayed_packets"] == 1

    asyncio.run(run())


def test_broadcast_presence_skips_broken_client():
    class BrokenWebSocket:
        async def send_json(self, message: dict) -> None:
            raise RuntimeError("Cannot write to closing transport")

    class HealthyWebSocket:
        def __init__(self) -> None:
            self.messages = []

        async def send_json(self, message: dict) -> None:
            self.messages.append(message)

    async def run() -> None:
        relay = WebIntercomServer("secret")
        broken_ws = BrokenWebSocket()
        healthy_ws = HealthyWebSocket()
        relay.clients[broken_ws] = Client(broken_ws, "client-0001", "bad", "main", time.monotonic())
        relay.clients[healthy_ws] = Client(healthy_ws, "client-0002", "good", "main", time.monotonic())

        await relay.broadcast_presence("main")

        assert healthy_ws.messages == [
            {
                "type": "presence",
                "room": "main",
                "count": 2,
                "clients": ["bad", "good"],
                "active_stream_ids": [],
            }
        ]

    asyncio.run(run())


def test_broadcast_presence_does_not_wait_for_slow_client():
    class SlowWebSocket:
        async def send_json(self, message: dict) -> None:
            await asyncio.sleep(1)

    class HealthyWebSocket:
        def __init__(self) -> None:
            self.messages = []

        async def send_json(self, message: dict) -> None:
            self.messages.append(message)

    async def run() -> None:
        relay = WebIntercomServer("secret")
        slow_ws = SlowWebSocket()
        healthy_ws = HealthyWebSocket()
        relay.clients[slow_ws] = Client(slow_ws, "client-0001", "slow", "main", time.monotonic())
        relay.clients[healthy_ws] = Client(healthy_ws, "client-0002", "good", "main", time.monotonic())

        started_at = time.perf_counter()
        await relay.broadcast_presence("main")
        elapsed = time.perf_counter() - started_at

        assert elapsed < 0.5
        assert healthy_ws.messages[-1]["clients"] == ["good", "slow"]

    asyncio.run(run())


def test_broadcast_presence_includes_active_stream_ids():
    class HealthyWebSocket:
        def __init__(self) -> None:
            self.messages = []

        async def send_json(self, message: dict) -> None:
            self.messages.append(message)

    async def run() -> None:
        relay = WebIntercomServer("secret")
        websocket = HealthyWebSocket()
        client = Client(websocket, "client-0001", "alice", "main", time.monotonic())
        client.stream_ids.update({99, 7})
        relay.clients[websocket] = client

        await relay.broadcast_presence("main")

        assert websocket.messages[-1]["active_stream_ids"] == [7, 99]

    asyncio.run(run())


def test_e_model_degrades_with_delay_and_late_drop():
    good_r, good_mos = estimate_e_model(delay_ms=40, late_drop_percent=0)
    poor_r, poor_mos = estimate_e_model(delay_ms=250, late_drop_percent=5)

    assert good_r > poor_r
    assert good_mos > poor_mos
    assert 1 <= poor_mos <= 4.5


def test_client_late_drop_rate_uses_received_packets_denominator():
    derived = derive_client_qoe_metrics(
        {
            "received_packets": 10,
            "late_dropped_packets": 2,
            "playback_queue_seconds": 0,
            "estimated_owd_ms": 0,
            "callback_interval_mean_ms": 8,
        }
    )

    assert derived["late_drop_rate_percent"] == 20.0


def test_metrics_workbook_contains_processed_sheets(tmp_path):
    relay = WebIntercomServer("secret")
    joined_at = time.monotonic()
    relay.metrics.record_join("client-0001", "alice", "main", joined_at)
    relay.metrics.record_client_metrics(
        "client-0001",
        {
            "captured_frames": 10,
            "sent_packets": 10,
            "sent_bytes": 6600,
            "sent_payload_bytes": 6400,
            "received_packets": 8,
            "received_bytes": 5280,
            "received_payload_bytes": 5120,
            "played_packets": 8,
            "rtt_ms": 4,
            "estimated_owd_ms": 2,
            "rfc3550_jitter_ms": 0.8,
            "playback_queue_seconds": 0.02,
            "late_dropped_packets": 0,
            "buffer_underrun_events": 0,
            "session_duration_seconds": 5,
            "last_sent_kbps": 12.8,
            "last_rx_kbps": 10.2,
        },
    )
    relay.metrics.record_audio_in(660, 640)
    relay.metrics.record_audio_relay(660, 640)
    relay.metrics.sample([("client-0001", "alice", "main")])
    relay.metrics.sample([("client-0001", "alice", "main")])
    output = tmp_path / "web_metrics.xlsx"

    relay.metrics.write_workbook(output)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "QoS Summary",
        "QoE Summary",
        "Jitter CDF",
        "Client Summary",
        "Time Series",
        "Paper Metrics",
    ]
    assert workbook["Summary"]["A1"].value == "Secure Web Intercom Measurement Summary"
    assert workbook["Summary"].row_dimensions[1].height >= 36
    assert workbook["Summary"].freeze_panes == "A5"
    assert len(workbook["Summary"]._charts) == 0
    summary_labels = [cell.value for cell in workbook["Summary"]["A"]]
    assert "Audio received" not in summary_labels
    assert "Authentication failures" not in summary_labels
    assert "P95 playout latency" in summary_labels
    assert workbook["QoS Summary"]["A1"].value == "Application-Level QoS Summary"
    assert len(workbook["QoS Summary"]._charts) == 1
    assert workbook["QoE Summary"]["A1"].value == "QoE and E-Model Summary"
    assert workbook["Client Summary"]["A4"].value == "client_id"
    client_headers = [cell.value for cell in workbook["Client Summary"][4]]
    assert "malformed_audio_packets" not in client_headers
    assert "captured_frames" not in client_headers
    assert workbook["Time Series"]["A4"].value == "timestamp"
    sample_headers = [cell.value for cell in workbook["Time Series"][4]]
    assert "rx_audio_bytes" not in sample_headers
    assert "relayed_bytes" not in sample_headers
    assert "browser_sent_bytes" not in sample_headers
    assert "auth_failures" not in sample_headers
    assert workbook["Paper Metrics"]["A4"].value == "Metric group"


def test_qos_summary_late_drop_rate_uses_received_denominator(tmp_path):
    relay = WebIntercomServer("secret")
    relay.metrics.record_join("client-0001", "alice", "main", time.monotonic())
    relay.metrics.record_client_metrics(
        "client-0001",
        {
            "received_packets": 10,
            "late_dropped_packets": 2,
            "playback_queue_seconds": 0,
            "estimated_owd_ms": 0,
            "callback_interval_mean_ms": 8,
        },
    )
    relay.metrics.sample([("client-0001", "alice", "main")])
    output = tmp_path / "qos_drop_rate.xlsx"

    relay.metrics.write_workbook(output)

    sheet = load_workbook(output, data_only=True)["QoS Summary"]
    values = {row[0].value: row[1].value for row in sheet.iter_rows(min_row=5, max_col=2)}
    assert values["Late drop rate"] == 20


def test_async_workbook_writer_uses_picklable_snapshot(tmp_path):
    async def run() -> None:
        relay = WebIntercomServer("secret")
        relay.metrics.record_join("client-0001", "alice", "main", time.monotonic())
        relay.metrics.record_client_metrics(
            "client-0001",
            {
                "received_packets": 10,
                "late_dropped_packets": 1,
                "playback_queue_seconds": 0.01,
                "estimated_owd_ms": 2,
                "callback_interval_mean_ms": 8,
            },
        )
        relay.metrics.sample([("client-0001", "alice", "main")])
        output = tmp_path / "async_metrics.xlsx"

        try:
            await relay.write_metrics_workbook(str(output))
        finally:
            await relay.close()

        assert output.exists()

    asyncio.run(run())
