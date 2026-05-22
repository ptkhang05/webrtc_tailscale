"""Processed measurement workbook export for the web intercom server."""

from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import threading
import time
from typing import Mapping


MetricValue = int | float | str
MetricRow = dict[str, MetricValue]

E_MODEL_R0 = 93.2
E_MODEL_IE_PCM16 = 0.0
E_MODEL_BPL_NO_PLC = 10.0


SERVER_GUIDE = {
    "timestamp": ("Timestamp", "", "Local time when the measurement row was captured."),
    "uptime_seconds": ("Uptime", "seconds", "Elapsed server runtime when the sample was captured."),
    "active_clients": ("Active clients", "clients", "Connected browser clients at this sample."),
    "active_rooms": ("Active rooms", "rooms", "Rooms with at least one connected client."),
    "join_events": ("Join events", "events", "Successful browser client joins."),
    "leave_events": ("Leave events", "events", "Browser client disconnects."),
    "auth_failures": ("Authentication failures", "events", "Join attempts rejected because the room key was wrong."),
    "auth_rate_limit_rejections": ("Auth rate-limit rejections", "events", "Join attempts rejected because an IP exceeded the failed room-key attempt limit."),
    "connection_limit_rejections": ("Connection limit rejections", "events", "Connections rejected because server or per-IP connection limits were reached."),
    "invalid_messages": ("Invalid control messages", "events", "Malformed or unexpected text messages."),
    "rx_audio_packets": ("Audio packets received", "packets", "Binary audio packets accepted by the server."),
    "rx_audio_bytes": ("Audio network bytes received", "bytes", "Audio packet bytes received from browsers, including the application packet header."),
    "rx_audio_payload_bytes": ("Audio payload bytes received", "bytes", "PCM16 audio payload bytes received from browsers, excluding the application packet header."),
    "relayed_packets": ("Audio packets relayed", "packets", "Audio packets forwarded to other clients."),
    "relayed_bytes": ("Audio network bytes relayed", "bytes", "Audio packet bytes forwarded to browsers, including the application packet header."),
    "relayed_payload_bytes": ("Audio payload bytes relayed", "bytes", "PCM16 payload bytes forwarded to browsers, excluding the application packet header."),
    "relay_send_drops": ("Relay send drops", "packets", "Outbound audio packets dropped because a recipient WebSocket was slow or unavailable."),
    "dropped_audio_frames": ("Dropped audio frames", "frames", "Audio frames rejected because they were empty or too large."),
    "avg_rx_kbps": ("Average receive network bitrate", "kbps", "Average server audio receive rate since start, including the application packet header."),
    "avg_rx_payload_kbps": ("Average receive payload bitrate", "kbps", "Average server PCM16 payload receive rate since start."),
    "avg_relayed_kbps": ("Average relay network bitrate", "kbps", "Average server audio forwarding rate since start, including the application packet header."),
    "avg_relayed_payload_kbps": ("Average relay payload bitrate", "kbps", "Average server PCM16 payload forwarding rate since start."),
    "interval_rx_kbps": ("Interval receive bitrate", "kbps", "Audio receive bitrate during the latest interval."),
    "interval_rx_payload_kbps": ("Interval receive payload bitrate", "kbps", "PCM16 payload receive bitrate during the latest interval."),
    "interval_relayed_kbps": ("Interval relay bitrate", "kbps", "Audio forwarding bitrate during the latest interval."),
    "interval_relayed_payload_kbps": ("Interval relay payload bitrate", "kbps", "PCM16 payload forwarding bitrate during the latest interval."),
    "interval_rx_pps": ("Interval receive packet rate", "packets/s", "Audio packets received per second in the latest interval."),
    "interval_relayed_pps": ("Interval relay packet rate", "packets/s", "Audio packets relayed per second in the latest interval."),
    "browser_captured_frames": ("Browser captured frames", "frames", "Total microphone callback frames reported by active browsers."),
    "browser_sent_packets": ("Browser sent packets", "packets", "Total audio packets browsers report sending."),
    "browser_sent_bytes": ("Browser sent bytes", "bytes", "Total audio bytes browsers report sending."),
    "browser_sent_payload_bytes": ("Browser sent payload bytes", "bytes", "Total PCM16 payload bytes browsers report sending."),
    "browser_received_packets": ("Browser received packets", "packets", "Total audio packets browsers report receiving."),
    "browser_received_bytes": ("Browser received bytes", "bytes", "Total audio bytes browsers report receiving."),
    "browser_received_payload_bytes": ("Browser received payload bytes", "bytes", "Total PCM16 payload bytes browsers report receiving."),
    "browser_played_packets": ("Browser played packets", "packets", "Total received packets scheduled for speaker playback."),
    "browser_capture_errors": ("Browser capture errors", "events", "Microphone capture errors reported by browsers."),
    "browser_malformed_audio_packets": ("Malformed browser audio packets", "packets", "Audio packets rejected by browsers because the application header was invalid."),
    "browser_network_loss_packets": ("Network sequence gaps", "packets", "Estimated missing audio packets from stream sequence discontinuities."),
    "browser_late_dropped_packets": ("Late dropped packets", "packets", "Packets browsers actually dropped because the per-stream playback queue budget was exceeded."),
    "browser_queue_overflow_dropped_packets": ("Queue overflow drops", "packets", "Packets dropped to prevent WebSocket/TCP accumulated delay from growing the playback queue."),
    "browser_buffer_underrun_events": ("Buffer underrun events", "events", "Times browser playback found the playout schedule had already run dry."),
    "browser_buffer_underrun_seconds": ("Buffer underrun duration", "seconds", "Total browser playback gap duration reported by clients."),
    "browser_max_buffer_underrun_ms": ("Max buffer underrun", "ms", "Largest single browser playback gap duration."),
    "browser_avg_rtt_ms": ("Average RTT", "ms", "Average browser-to-server WebSocket control round-trip time."),
    "browser_max_rtt_ms": ("Max RTT", "ms", "Largest browser-to-server WebSocket control round-trip time."),
    "browser_avg_estimated_owd_ms": ("Average estimated OWD", "ms", "Average one-way delay estimate as RTT/2. This does not require synchronized clocks."),
    "browser_max_jitter_ms": ("Max RFC3550 jitter", "ms", "Maximum browser-estimated inter-arrival jitter using the RFC 3550 estimator."),
    "browser_avg_playout_latency_ms": ("Average playout latency estimate", "ms", "Estimated client-side network plus playback queue delay."),
    "browser_avg_estimated_mos": ("Average estimated MOS", "MOS", "Average objective MOS estimate from the simplified ITU-T G.107 E-model."),
    "browser_min_estimated_mos": ("Minimum estimated MOS", "MOS", "Worst client MOS estimate at this sample."),
    "browser_max_callback_stddev_ms": ("Max callback stddev", "ms", "Largest browser AudioWorklet callback interval standard deviation."),
    "browser_resampled_frames": ("Browser resampled frames", "frames", "Captured frames resampled to 16 kHz before network transmission."),
    "browser_capture_queue_dropped_frames": ("Capture queue drops", "frames", "Captured frames dropped because the client-side send/resampling queue was full."),
    "browser_max_audio_context_sample_rate": ("Max audio context sample rate", "Hz", "Largest actual browser AudioContext sample rate reported by connected clients."),
    "browser_max_active_remote_streams": ("Max active remote streams", "streams", "Largest remote stream count reported by connected clients."),
}


CLIENT_GUIDE = {
    "client_id": ("Client ID", "", "Server-assigned unique client connection ID."),
    "name": ("Name", "", "Browser display name."),
    "room": ("Room", "", "Intercom room name."),
    "status": ("Status", "", "Whether the client is active at workbook export time."),
    "session_duration_seconds": ("Session duration", "seconds", "Elapsed browser session time reported by the client."),
    "uptime_seconds": ("Client uptime", "seconds", "Elapsed time since this client joined."),
    "captured_frames": ("Captured frames", "frames", "Microphone callback frames captured in the browser."),
    "sent_packets": ("Sent packets", "packets", "Audio packets sent from the browser to the server."),
    "sent_bytes": ("Sent network bytes", "bytes", "Audio packet bytes sent from the browser to the server, including the application packet header."),
    "sent_payload_bytes": ("Sent payload bytes", "bytes", "PCM16 payload bytes sent from the browser to the server."),
    "received_packets": ("Received packets", "packets", "Audio packets received from the server."),
    "received_bytes": ("Received network bytes", "bytes", "Audio packet bytes received from the server, including the application packet header."),
    "received_payload_bytes": ("Received payload bytes", "bytes", "PCM16 payload bytes received from the server."),
    "played_packets": ("Played packets", "packets", "Received packets scheduled for speaker playback."),
    "capture_errors": ("Capture errors", "events", "Browser-side capture/send errors."),
    "malformed_audio_packets": ("Malformed audio packets", "packets", "Received binary messages rejected by the browser because their application header was invalid."),
    "network_loss_packets": ("Network sequence gaps", "packets", "Estimated missing audio packets from sequence discontinuities; this is not counted as a late drop unless playout discards a packet."),
    "late_dropped_packets": ("Late dropped packets", "packets", "Received packets actually dropped because the per-stream playback queue budget was exceeded."),
    "queue_overflow_dropped_packets": ("Queue overflow drops", "packets", "Packets dropped to cap accumulated playback queue delay."),
    "buffer_underrun_events": ("Buffer underrun events", "events", "Times playback discovered that the scheduled playout buffer had run dry."),
    "buffer_underrun_seconds": ("Buffer underrun duration", "seconds", "Total playback gap duration caused by buffer underruns."),
    "max_buffer_underrun_ms": ("Max buffer underrun", "ms", "Largest single playback gap caused by an underrun."),
    "rtt_ms": ("WebSocket RTT", "ms", "Browser-to-server control round-trip time measured with a lightweight QoS ping."),
    "estimated_owd_ms": ("Estimated one-way delay", "ms", "Approximate browser-to-server one-way delay, computed as RTT/2 under symmetric-path assumption."),
    "rfc3550_jitter_ms": ("RFC3550 jitter", "ms", "Inter-arrival jitter estimate computed from audio packet send and receive timestamp differences."),
    "callback_interval_mean_ms": ("AudioWorklet callback mean", "ms", "Mean interval between AudioWorklet process callbacks in audio time."),
    "callback_interval_stddev_ms": ("AudioWorklet callback stddev", "ms", "Standard deviation of AudioWorklet process callback intervals."),
    "callback_interval_max_ms": ("AudioWorklet callback max", "ms", "Maximum observed AudioWorklet process callback interval."),
    "worklet_message_interval_mean_ms": ("Worklet message mean", "ms", "Mean interval between AudioWorklet messages received by the main thread."),
    "worklet_message_interval_stddev_ms": ("Worklet message stddev", "ms", "Standard deviation of AudioWorklet-to-main-thread message intervals."),
    "worklet_message_interval_max_ms": ("Worklet message max", "ms", "Maximum observed AudioWorklet-to-main-thread message interval."),
    "audio_context_sample_rate": ("Audio context sample rate", "Hz", "Actual browser AudioContext sample rate; may differ from the requested 16 kHz."),
    "resampled_frames": ("Resampled frames", "frames", "Captured frames resampled to 16 kHz before transmission."),
    "capture_queue_dropped_frames": ("Capture queue drops", "frames", "Captured frames dropped because the send/resampling queue was full."),
    "active_remote_streams": ("Active remote streams", "streams", "Remote stream states currently tracked by the browser."),
    "last_sent_kbps": ("Latest send bitrate", "kbps", "Browser-calculated send bitrate in the latest UI interval."),
    "last_rx_kbps": ("Latest receive bitrate", "kbps", "Browser-calculated receive bitrate in the latest UI interval."),
    "playback_queue_seconds": ("Playback queue", "seconds", "Estimated browser playback buffer delay."),
    "estimated_playout_latency_ms": ("Estimated playout latency", "ms", "Estimated one-way network delay plus current playback queue delay and capture quantum."),
    "late_drop_rate_percent": ("Late drop rate", "%", "Late dropped packets divided by received packets."),
    "buffer_underrun_rate_per_min": ("Buffer underrun rate", "events/min", "Playback underrun events normalized by browser session duration."),
    "r_factor": ("E-model R-factor", "score", "Simplified ITU-T G.107 R-factor estimated from delay and late/drop impairment."),
    "estimated_mos": ("Estimated MOS", "MOS", "Estimated Mean Opinion Score mapped from the R-factor."),
    "mos_quality": ("MOS quality", "", "Plain-language quality descriptor derived from estimated MOS."),
}


@dataclass
class ClientMetricState:
    client_id: str
    name: str
    room: str
    joined_at: float
    active: bool = True
    latest: dict[str, MetricValue] = field(default_factory=dict)
    samples: list[MetricRow] = field(default_factory=list)


class WebIntercomMetrics:
    def __init__(self) -> None:
        self.started_at = time.monotonic()
        self._lock = threading.Lock()
        self._next_client_id = 1
        self.server_counters: defaultdict[str, int] = defaultdict(int)
        self.client_states: dict[str, ClientMetricState] = {}
        self.server_samples: list[MetricRow] = []
        self._last_server_sample: MetricRow | None = None

    def next_client_id(self) -> str:
        with self._lock:
            value = f"client-{self._next_client_id:04d}"
            self._next_client_id += 1
            return value

    def inc(self, field: str, amount: int = 1) -> None:
        if amount < 0:
            raise ValueError("metric increment cannot be negative")
        with self._lock:
            self.server_counters[field] += amount

    def record_join(self, client_id: str, name: str, room: str, joined_at: float) -> None:
        with self._lock:
            self.server_counters["join_events"] += 1
            self.client_states[client_id] = ClientMetricState(
                client_id=client_id,
                name=name,
                room=room,
                joined_at=joined_at,
            )

    def record_leave(self, client_id: str) -> None:
        with self._lock:
            self.server_counters["leave_events"] += 1
            if client_id in self.client_states:
                self.client_states[client_id].active = False

    def record_audio_in(self, byte_count: int, payload_byte_count: int | None = None) -> None:
        with self._lock:
            self.server_counters["rx_audio_packets"] += 1
            self.server_counters["rx_audio_bytes"] += byte_count
            self.server_counters["rx_audio_payload_bytes"] += payload_byte_count if payload_byte_count is not None else byte_count

    def record_audio_relay(self, byte_count: int, payload_byte_count: int | None = None) -> None:
        with self._lock:
            self.server_counters["relayed_packets"] += 1
            self.server_counters["relayed_bytes"] += byte_count
            self.server_counters["relayed_payload_bytes"] += payload_byte_count if payload_byte_count is not None else byte_count

    def record_client_metrics(self, client_id: str, metrics: Mapping[str, object]) -> None:
        with self._lock:
            state = self.client_states.get(client_id)
            if state is None:
                return
            cleaned = {
                field: _clean_metric_value(value)
                for field, value in metrics.items()
                if field in CLIENT_GUIDE
            }
            merged = {**state.latest, **cleaned}
            cleaned.update(derive_client_qoe_metrics(merged))
            state.latest.update(cleaned)
            state.samples.append(
                {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "client_id": client_id,
                    "name": state.name,
                    "room": state.room,
                    **cleaned,
                }
            )

    def sample(self, active_clients: list[tuple[str, str, str]]) -> MetricRow:
        with self._lock:
            uptime = time.monotonic() - self.started_at
            counters = dict(self.server_counters)
            active_ids = {client_id for client_id, _name, _room in active_clients}
            active_rooms = {room for _client_id, _name, room in active_clients}
            browser_totals = self._browser_totals(active_ids)
            row: MetricRow = {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "uptime_seconds": round(uptime, 3),
                "active_clients": len(active_ids),
                "active_rooms": len(active_rooms),
                "join_events": counters.get("join_events", 0),
                "leave_events": counters.get("leave_events", 0),
                "auth_failures": counters.get("auth_failures", 0),
                "auth_rate_limit_rejections": counters.get("auth_rate_limit_rejections", 0),
                "connection_limit_rejections": counters.get("connection_limit_rejections", 0),
                "invalid_messages": counters.get("invalid_messages", 0),
                "rx_audio_packets": counters.get("rx_audio_packets", 0),
                "rx_audio_bytes": counters.get("rx_audio_bytes", 0),
                "rx_audio_payload_bytes": counters.get("rx_audio_payload_bytes", 0),
                "relayed_packets": counters.get("relayed_packets", 0),
                "relayed_bytes": counters.get("relayed_bytes", 0),
                "relayed_payload_bytes": counters.get("relayed_payload_bytes", 0),
                "relay_send_drops": counters.get("relay_send_drops", 0),
                "dropped_audio_frames": counters.get("dropped_audio_frames", 0),
                "avg_rx_kbps": kbps(counters.get("rx_audio_bytes", 0), uptime),
                "avg_rx_payload_kbps": kbps(counters.get("rx_audio_payload_bytes", 0), uptime),
                "avg_relayed_kbps": kbps(counters.get("relayed_bytes", 0), uptime),
                "avg_relayed_payload_kbps": kbps(counters.get("relayed_payload_bytes", 0), uptime),
                **browser_totals,
            }
            self._add_interval_fields(row)
            self.server_samples.append(row)
            self._last_server_sample = row
            return dict(row)

    def snapshot(self) -> tuple[list[MetricRow], list[ClientMetricState]]:
        with self._lock:
            return deepcopy(self.server_samples), deepcopy(list(self.client_states.values()))

    def write_workbook(self, path: str | Path) -> None:
        server_samples, client_states = self.snapshot()
        write_metrics_workbook(Path(path), server_samples, client_states)

    def _browser_totals(self, active_ids: set[str]) -> dict[str, int]:
        totals = {
            "browser_captured_frames": 0,
            "browser_sent_packets": 0,
            "browser_sent_bytes": 0,
            "browser_sent_payload_bytes": 0,
            "browser_received_packets": 0,
            "browser_received_bytes": 0,
            "browser_received_payload_bytes": 0,
            "browser_played_packets": 0,
            "browser_capture_errors": 0,
            "browser_malformed_audio_packets": 0,
            "browser_network_loss_packets": 0,
            "browser_late_dropped_packets": 0,
            "browser_queue_overflow_dropped_packets": 0,
            "browser_buffer_underrun_events": 0,
            "browser_buffer_underrun_seconds": 0.0,
            "browser_max_buffer_underrun_ms": 0.0,
            "browser_avg_rtt_ms": 0.0,
            "browser_max_rtt_ms": 0.0,
            "browser_avg_estimated_owd_ms": 0.0,
            "browser_max_jitter_ms": 0.0,
            "browser_avg_playout_latency_ms": 0.0,
            "browser_avg_estimated_mos": 0.0,
            "browser_min_estimated_mos": 0.0,
            "browser_max_callback_stddev_ms": 0.0,
            "browser_resampled_frames": 0,
            "browser_capture_queue_dropped_frames": 0,
            "browser_max_audio_context_sample_rate": 0.0,
            "browser_max_active_remote_streams": 0,
        }
        mapping = {
            "captured_frames": "browser_captured_frames",
            "sent_packets": "browser_sent_packets",
            "sent_bytes": "browser_sent_bytes",
            "sent_payload_bytes": "browser_sent_payload_bytes",
            "received_packets": "browser_received_packets",
            "received_bytes": "browser_received_bytes",
            "received_payload_bytes": "browser_received_payload_bytes",
            "played_packets": "browser_played_packets",
            "capture_errors": "browser_capture_errors",
            "malformed_audio_packets": "browser_malformed_audio_packets",
            "network_loss_packets": "browser_network_loss_packets",
            "late_dropped_packets": "browser_late_dropped_packets",
            "queue_overflow_dropped_packets": "browser_queue_overflow_dropped_packets",
            "buffer_underrun_events": "browser_buffer_underrun_events",
            "resampled_frames": "browser_resampled_frames",
            "capture_queue_dropped_frames": "browser_capture_queue_dropped_frames",
        }
        rtt_values: list[float] = []
        owd_values: list[float] = []
        jitter_values: list[float] = []
        latency_values: list[float] = []
        mos_values: list[float] = []
        callback_stddev_values: list[float] = []
        for client_id in active_ids:
            state = self.client_states.get(client_id)
            if state is None:
                continue
            for source, target in mapping.items():
                totals[target] += int(_as_float(state.latest.get(source)))
            totals["browser_buffer_underrun_seconds"] += _as_float(state.latest.get("buffer_underrun_seconds"))
            totals["browser_max_buffer_underrun_ms"] = max(
                _as_float(totals["browser_max_buffer_underrun_ms"]),
                _as_float(state.latest.get("max_buffer_underrun_ms")),
            )
            _append_positive(rtt_values, state.latest.get("rtt_ms"))
            _append_positive(owd_values, state.latest.get("estimated_owd_ms"))
            _append_positive(jitter_values, state.latest.get("rfc3550_jitter_ms"))
            _append_positive(latency_values, state.latest.get("estimated_playout_latency_ms"))
            _append_positive(mos_values, state.latest.get("estimated_mos"))
            _append_positive(callback_stddev_values, state.latest.get("callback_interval_stddev_ms"))
            totals["browser_max_audio_context_sample_rate"] = max(
                _as_float(totals["browser_max_audio_context_sample_rate"]),
                _as_float(state.latest.get("audio_context_sample_rate")),
            )
            totals["browser_max_active_remote_streams"] = max(
                _as_int(totals["browser_max_active_remote_streams"]),
                _as_int(state.latest.get("active_remote_streams")),
            )
        totals["browser_avg_rtt_ms"] = _mean(rtt_values)
        totals["browser_max_rtt_ms"] = _max_values(rtt_values)
        totals["browser_avg_estimated_owd_ms"] = _mean(owd_values)
        totals["browser_max_jitter_ms"] = _max_values(jitter_values)
        totals["browser_avg_playout_latency_ms"] = _mean(latency_values)
        totals["browser_avg_estimated_mos"] = _mean(mos_values)
        totals["browser_min_estimated_mos"] = _min_values(mos_values)
        totals["browser_max_callback_stddev_ms"] = _max_values(callback_stddev_values)
        totals["browser_buffer_underrun_seconds"] = round(_as_float(totals["browser_buffer_underrun_seconds"]), 3)
        return totals

    def _add_interval_fields(self, row: MetricRow) -> None:
        previous = self._last_server_sample
        elapsed = _as_float(row["uptime_seconds"]) - _as_float(previous.get("uptime_seconds")) if previous else 0.0
        intervals = {
            "rx_audio_bytes": ("interval_rx_kbps", 8 / 1000),
            "rx_audio_payload_bytes": ("interval_rx_payload_kbps", 8 / 1000),
            "relayed_bytes": ("interval_relayed_kbps", 8 / 1000),
            "relayed_payload_bytes": ("interval_relayed_payload_kbps", 8 / 1000),
            "rx_audio_packets": ("interval_rx_pps", 1),
            "relayed_packets": ("interval_relayed_pps", 1),
        }
        for source, (target, scale) in intervals.items():
            if previous is None or elapsed <= 0:
                row[target] = 0.0
                continue
            delta = max(0.0, _as_float(row.get(source)) - _as_float(previous.get(source)))
            row[target] = round(delta * scale / elapsed, 3)


def kbps(byte_count: int | float, uptime_seconds: int | float) -> float:
    uptime = _as_float(uptime_seconds)
    if uptime <= 0:
        return 0.0
    return round((_as_float(byte_count) * 8) / uptime / 1000, 3)


def derive_client_qoe_metrics(metrics: Mapping[str, object]) -> MetricRow:
    playback_queue_ms = _as_float(metrics.get("playback_queue_seconds")) * 1000
    estimated_owd_ms = _as_float(metrics.get("estimated_owd_ms"))
    capture_quantum_ms = _as_float(metrics.get("callback_interval_mean_ms")) or 8.0
    estimated_playout_latency_ms = max(0.0, capture_quantum_ms + estimated_owd_ms + playback_queue_ms)

    late_drops = _as_float(metrics.get("late_dropped_packets"))
    received_packets = _as_float(metrics.get("received_packets"))
    late_drop_rate_percent = percentage(late_drops, received_packets)

    session_duration_seconds = _as_float(metrics.get("session_duration_seconds"))
    underrun_rate_per_min = 0.0
    if session_duration_seconds > 0:
        underrun_rate_per_min = _as_float(metrics.get("buffer_underrun_events")) / (session_duration_seconds / 60)

    r_factor, estimated_mos = estimate_e_model(
        delay_ms=estimated_playout_latency_ms,
        late_drop_percent=late_drop_rate_percent,
    )
    return {
        "estimated_playout_latency_ms": round(estimated_playout_latency_ms, 3),
        "late_drop_rate_percent": round(late_drop_rate_percent, 3),
        "buffer_underrun_rate_per_min": round(underrun_rate_per_min, 3),
        "r_factor": round(r_factor, 3),
        "estimated_mos": round(estimated_mos, 3),
        "mos_quality": mos_quality_label(estimated_mos),
    }


def estimate_e_model(delay_ms: float, late_drop_percent: float) -> tuple[float, float]:
    delay = max(0.0, delay_ms)
    delay_impairment = 0.024 * delay
    if delay > 177.3:
        delay_impairment += 0.11 * (delay - 177.3)

    packet_impairment = E_MODEL_IE_PCM16
    loss = max(0.0, late_drop_percent)
    if loss > 0:
        packet_impairment += (95.0 - E_MODEL_IE_PCM16) * loss / (loss + E_MODEL_BPL_NO_PLC)

    r_factor = max(0.0, min(100.0, E_MODEL_R0 - delay_impairment - packet_impairment))
    return r_factor, r_factor_to_mos(r_factor)


def r_factor_to_mos(r_factor: float) -> float:
    r = max(0.0, min(100.0, _as_float(r_factor)))
    if r <= 0:
        return 1.0
    if r >= 100:
        return 4.5
    return round(1 + 0.035 * r + 7e-6 * r * (r - 60) * (100 - r), 3)


def mos_quality_label(mos: float) -> str:
    value = _as_float(mos)
    if value >= 4.3:
        return "Excellent"
    if value >= 4.0:
        return "Good"
    if value >= 3.6:
        return "Fair"
    if value >= 3.1:
        return "Poor"
    if value >= 2.6:
        return "Bad"
    return "Not recommended"


def percentage(numerator: object, denominator: object) -> float:
    value = _as_float(denominator)
    if value <= 0:
        return 0.0
    return (_as_float(numerator) / value) * 100


def write_metrics_workbook(
    path: Path,
    server_samples: list[MetricRow],
    client_states: list[ClientMetricState],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl. Run: python -m pip install -r requirements.txt") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    qos_summary = workbook.create_sheet("QoS Summary")
    qoe_summary = workbook.create_sheet("QoE Summary")
    jitter_cdf = workbook.create_sheet("Jitter CDF")
    client_summary = workbook.create_sheet("Client Summary")
    time_series = workbook.create_sheet("Time Series")
    paper_metrics = workbook.create_sheet("Paper Metrics")

    latest = server_samples[-1] if server_samples else {}
    _build_summary(summary, latest, server_samples, client_states, Font, PatternFill)
    _build_qos_summary(qos_summary, latest, server_samples, client_states, Font, PatternFill)
    _build_qoe_summary(qoe_summary, latest, client_states, Font, PatternFill)
    _build_jitter_cdf(jitter_cdf, client_states, LineChart, Reference, Font, PatternFill)
    _build_client_summary(client_summary, client_states, Table, TableStyleInfo, Font, PatternFill)
    _build_time_series(time_series, server_samples, Table, TableStyleInfo, Font, PatternFill)
    _build_paper_metrics(paper_metrics, Font, PatternFill)
    _add_qos_chart(qos_summary, server_samples, LineChart, Reference)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A5"
        _size_columns(sheet, Alignment)

    temp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    workbook.save(temp_path)
    temp_path.replace(path)


def _build_summary(sheet: object, latest: MetricRow, rows: list[MetricRow], clients: list[ClientMetricState], font_cls: object, fill_cls: object) -> None:
    _title(sheet, "Secure Web Intercom Measurement Summary", font_cls, fill_cls)
    headers = ["Metric", "Value", "Unit", "Interpretation"]
    _write_row(sheet, 4, headers)
    _style_header(sheet, "A4:D4", font_cls, fill_cls)

    received = _as_float(latest.get("browser_received_packets"))
    relayed = _as_float(latest.get("relayed_packets"))
    jitter_values = _client_sample_values(clients, "rfc3550_jitter_ms")
    latency_values = _client_sample_values(clients, "estimated_playout_latency_ms")
    owd_values = _client_sample_values(clients, "estimated_owd_ms")
    callback_stddev_values = _client_sample_values(clients, "callback_interval_stddev_ms")

    summary_rows = [
        ("Measurement duration", _as_float(latest.get("uptime_seconds")), "seconds", "How long this measurement run has been active."),
        ("Sample count", len(rows), "samples", "Number of server-side measurement intervals saved."),
        ("Active clients", _as_int(latest.get("active_clients")), "clients", "Browsers connected at the latest sample."),
        ("Total clients seen", len(clients), "clients", "Unique browser connections during the run."),
        ("Mean estimated OWD", _mean(owd_values) or _as_float(latest.get("browser_avg_estimated_owd_ms")), "ms", "RTT/2 LAN delay estimate used when client clocks are not synchronized."),
        ("P95 playout latency", _percentile(latency_values, 95), "ms", "High-percentile delay estimate for conversational QoE analysis."),
        ("P95 RFC3550 jitter", _percentile(jitter_values, 95), "ms", "High-percentile inter-arrival jitter for paper plots."),
        ("Network sequence gap rate", percentage(latest.get("browser_network_loss_packets"), received), "%", "Estimated sequence gaps divided by browser-received packets."),
        ("Late drop rate", percentage(latest.get("browser_late_dropped_packets"), received), "%", "Packets actually discarded because the playout queue budget was exceeded."),
        ("Buffer underrun events", _as_int(latest.get("browser_buffer_underrun_events")), "events", "Browser playout starvation events during the run."),
        ("Buffer underrun duration", _as_float(latest.get("browser_buffer_underrun_seconds")), "seconds", "Total estimated audible gap duration reported by browsers."),
        ("Average estimated MOS", _as_float(latest.get("browser_avg_estimated_mos")), "MOS", "Objective QoE estimate from the simplified E-model."),
        ("Minimum estimated MOS", _as_float(latest.get("browser_min_estimated_mos")), "MOS", "Worst current client MOS estimate."),
        ("Average payload relay bitrate", _as_float(latest.get("avg_relayed_payload_kbps")), "kbps", "Average forwarded PCM16 payload rate; useful for bandwidth discussion."),
        ("Peak payload relay bitrate", _max(rows, "interval_relayed_payload_kbps"), "kbps", "Highest interval forwarding rate observed by the server."),
        ("Relay send drop rate", percentage(latest.get("relay_send_drops"), relayed), "%", "Server-side drops caused by bounded relay queues or send timeouts."),
        ("AudioWorklet callback stddev max", _max_values(callback_stddev_values), "ms", "Worst browser capture callback timing instability."),
        ("Resampled frames", _as_int(latest.get("browser_resampled_frames")), "frames", "Captured frames converted to 16 kHz when hardware sample rate differed."),
        ("Capture queue drops", _as_int(latest.get("browser_capture_queue_dropped_frames")), "frames", "Captured frames discarded before send due to client-side queue pressure."),
    ]
    for index, row in enumerate(summary_rows, start=5):
        _write_row(sheet, index, row)
        if isinstance(row[1], float):
            sheet.cell(index, 2).number_format = "0.000"


def _build_qos_summary(sheet: object, latest: MetricRow, rows: list[MetricRow], clients: list[ClientMetricState], font_cls: object, fill_cls: object) -> None:
    _title(sheet, "Application-Level QoS Summary", font_cls, fill_cls)
    headers = ["Metric", "Value", "Unit", "Interpretation"]
    _write_row(sheet, 4, headers)
    _style_header(sheet, "A4:D4", font_cls, fill_cls)

    jitter_values = _client_sample_values(clients, "rfc3550_jitter_ms")
    latency_values = _client_sample_values(clients, "estimated_playout_latency_ms")
    owd_values = _client_sample_values(clients, "estimated_owd_ms")
    callback_stddev_values = _client_sample_values(clients, "callback_interval_stddev_ms")
    late_drops = _as_float(latest.get("browser_late_dropped_packets"))
    received = _as_float(latest.get("browser_received_packets"))
    network_gaps = _as_float(latest.get("browser_network_loss_packets"))
    relayed = _as_float(latest.get("relayed_packets"))

    qos_rows = [
        ("Estimated OWD mean", _mean(owd_values), "ms", "RTT/2 approximation for the browser-to-server path; useful without synchronized LAN clocks."),
        ("Estimated OWD p95", _percentile(owd_values, 95), "ms", "95th percentile one-way delay estimate across client metric samples."),
        ("RFC3550 jitter mean", _mean(jitter_values), "ms", "Mean inter-arrival jitter calculated from packet timestamp spacing."),
        ("RFC3550 jitter p95", _percentile(jitter_values, 95), "ms", "95th percentile jitter; this is more useful for paper plots than total byte counters."),
        ("RFC3550 jitter max", _max_values(jitter_values), "ms", "Worst observed browser jitter estimate."),
        ("Network sequence gaps", network_gaps, "packets", "Estimated missing packets from stream sequence discontinuities; separate from actual playout drops."),
        ("Network sequence gap rate", percentage(network_gaps, received), "%", "Sequence gaps divided by browser-received packets; useful when comparing Wi-Fi conditions."),
        ("Late drop rate", percentage(late_drops, received), "%", "Packets actually dropped because TCP/WebSocket delivery caused a per-stream queue budget overflow."),
        ("Buffer underrun events", _as_int(latest.get("browser_buffer_underrun_events")), "events", "Playback starvation events reported by browsers."),
        ("Buffer underrun duration", _as_float(latest.get("browser_buffer_underrun_seconds")), "seconds", "Cumulative audible gap duration from browser playout underruns."),
        ("Estimated playout latency p95", _percentile(latency_values, 95), "ms", "Approximate client playout path delay including RTT/2 and playback queue."),
        ("AudioWorklet callback stddev max", _max_values(callback_stddev_values), "ms", "Worst callback interval instability reported by browsers."),
        ("Relay send drop rate", percentage(latest.get("relay_send_drops"), relayed), "%", "Server relay drops divided by relayed packets; captures bounded-queue pressure."),
        ("Peak relay payload bitrate", _max(rows, "interval_relayed_payload_kbps"), "kbps", "Peak forwarded PCM16 payload rate during a measurement interval."),
    ]
    for index, row in enumerate(qos_rows, start=5):
        _write_row(sheet, index, row)
        if isinstance(row[1], float):
            sheet.cell(index, 2).number_format = "0.000"


def _build_qoe_summary(sheet: object, latest: MetricRow, clients: list[ClientMetricState], font_cls: object, fill_cls: object) -> None:
    _title(sheet, "QoE and E-Model Summary", font_cls, fill_cls)
    headers = ["Metric", "Value", "Unit", "Interpretation"]
    _write_row(sheet, 4, headers)
    _style_header(sheet, "A4:D4", font_cls, fill_cls)

    mos_values = _client_sample_values(clients, "estimated_mos")
    r_values = _client_sample_values(clients, "r_factor")
    latency_values = _client_sample_values(clients, "estimated_playout_latency_ms")
    latest_mos = _as_float(latest.get("browser_avg_estimated_mos"))
    latest_min_mos = _as_float(latest.get("browser_min_estimated_mos"))
    quality_class = mos_quality_label(latest_mos) if latest_mos > 0 else "No data"
    threshold_status = "No data" if latest_mos == 0 and not mos_values else ("OK" if latest_mos >= 3.6 else "Review")
    qoe_rows = [
        ("Average estimated MOS", _mean(mos_values) or latest_mos, "MOS", "Objective voice quality estimate mapped from the simplified ITU-T G.107 R-factor."),
        ("Minimum estimated MOS", _min_values(mos_values) or latest_min_mos, "MOS", "Worst observed client quality estimate."),
        ("Average R-factor", _mean(r_values), "score", "Transmission rating factor before conversion to MOS."),
        ("Minimum R-factor", _min_values(r_values), "score", "Worst observed R-factor."),
        ("Latency p95", _percentile(latency_values, 95), "ms", "High-percentile playout delay used for QoE interpretation."),
        ("Latest quality class", quality_class, "", "Quality descriptor for the latest average MOS sample."),
        ("Target threshold", 3.6, "MOS", "Fair quality threshold commonly used as a minimum acceptable intercom target."),
        ("Threshold status", threshold_status, "", "Review if estimated MOS is below the fair-quality target."),
    ]
    for index, row in enumerate(qoe_rows, start=5):
        _write_row(sheet, index, row)
        if isinstance(row[1], float):
            sheet.cell(index, 2).number_format = "0.000"

    _write_row(sheet, 16, ["Model note"])
    sheet["A16"].font = font_cls(bold=True)
    sheet["A17"] = (
        "The MOS value is an objective estimate, not a subjective listening test. "
        "It uses measured delay, late packet drops, and a simplified ITU-T G.107 E-model mapping."
    )
    sheet.merge_cells("A17:D18")


def _build_jitter_cdf(sheet: object, clients: list[ClientMetricState], chart_cls: object, reference_cls: object, font_cls: object, fill_cls: object) -> None:
    _title(sheet, "RFC3550 Jitter CDF", font_cls, fill_cls)
    fields = ["Percentile", "Jitter_ms"]
    _write_row(sheet, 4, fields)
    _style_header(sheet, "A4:B4", font_cls, fill_cls)
    values = _client_sample_values(clients, "rfc3550_jitter_ms")
    if not values:
        _write_row(sheet, 5, [0, 0.0])
        return

    percentiles = [0, 5, 10, 25, 50, 75, 90, 95, 99, 100]
    for row_index, percentile in enumerate(percentiles, start=5):
        _write_row(sheet, row_index, [percentile, _percentile(values, percentile)])

    chart = chart_cls()
    _format_line_chart(chart, "Jitter CDF", "Jitter (ms)", "Percentile")
    data = reference_cls(sheet, min_col=2, max_col=2, min_row=4, max_row=len(percentiles) + 4)
    chart.add_data(data, titles_from_data=True)
    categories = reference_cls(sheet, min_col=1, min_row=5, max_row=len(percentiles) + 4)
    chart.set_categories(categories)
    _style_chart_series(chart, ["4472C4"])
    sheet.add_chart(chart, "D5")


def _build_client_summary(sheet: object, clients: list[ClientMetricState], table_cls: object, style_cls: object, font_cls: object, fill_cls: object) -> None:
    _title(sheet, "Client Summary", font_cls, fill_cls)
    fields = [
        "client_id",
        "name",
        "room",
        "status",
        "session_duration_seconds",
        "sent_packets",
        "received_packets",
        "played_packets",
        "capture_errors",
        "network_loss_packets",
        "late_dropped_packets",
        "late_drop_rate_percent",
        "buffer_underrun_events",
        "buffer_underrun_seconds",
        "max_buffer_underrun_ms",
        "rtt_ms",
        "estimated_owd_ms",
        "rfc3550_jitter_ms",
        "estimated_playout_latency_ms",
        "buffer_underrun_rate_per_min",
        "r_factor",
        "estimated_mos",
        "mos_quality",
        "callback_interval_stddev_ms",
        "audio_context_sample_rate",
        "resampled_frames",
        "capture_queue_dropped_frames",
        "active_remote_streams",
    ]
    _write_table(sheet, fields, [_client_summary_row(client) for client in clients], table_cls, style_cls, font_cls, fill_cls)


def _build_time_series(sheet: object, rows: list[MetricRow], table_cls: object, style_cls: object, font_cls: object, fill_cls: object) -> None:
    _title(sheet, "Paper-Ready Time Series", font_cls, fill_cls)
    fields = [
        "timestamp",
        "uptime_seconds",
        "active_clients",
        "active_rooms",
        "interval_rx_payload_kbps",
        "interval_relayed_payload_kbps",
        "interval_rx_pps",
        "interval_relayed_pps",
        "browser_received_packets",
        "browser_network_loss_packets",
        "browser_late_dropped_packets",
        "browser_buffer_underrun_events",
        "browser_buffer_underrun_seconds",
        "browser_avg_rtt_ms",
        "browser_avg_estimated_owd_ms",
        "browser_max_jitter_ms",
        "browser_avg_playout_latency_ms",
        "browser_avg_estimated_mos",
        "browser_min_estimated_mos",
        "browser_max_callback_stddev_ms",
        "relay_send_drops",
        "dropped_audio_frames",
        "browser_resampled_frames",
        "browser_capture_queue_dropped_frames",
    ]
    _write_table(sheet, fields, rows, table_cls, style_cls, font_cls, fill_cls)


def _build_paper_metrics(sheet: object, font_cls: object, fill_cls: object) -> None:
    _title(sheet, "Paper Metric Selection", font_cls, fill_cls)
    fields = ["Metric group", "Exported metric", "Why it matters for the paper", "Use in Experimental Results"]
    _write_row(sheet, 4, fields)
    _style_header(sheet, "A4:D4", font_cls, fill_cls)
    rows = [
        (
            "Delay",
            "Mean/P95 estimated OWD and P95 playout latency",
            "Voice intercom quality is delay-sensitive; high percentiles expose TCP/WSS delay spikes better than averages alone.",
            "Report per network condition and compare against the conversational-delay target.",
        ),
        (
            "Jitter",
            "RFC3550 jitter mean/P95/max and Jitter CDF",
            "Inter-arrival jitter is a standard real-time media QoS indicator and directly supports distribution plots.",
            "Use the CDF sheet for the main paper figure.",
        ),
        (
            "Late delivery",
            "Late drop rate and network sequence gap rate",
            "WSS/TCP can hide transport loss but still creates packets that miss the playout budget.",
            "Report as impairment percentages under LAN, Wi-Fi, and congested Wi-Fi.",
        ),
        (
            "Playout continuity",
            "Buffer underrun events and duration",
            "Underruns are audible continuity failures and make the latency/jitter results interpretable.",
            "Use events/minute or total duration as supporting QoE evidence.",
        ),
        (
            "QoE",
            "R-factor, estimated MOS, and MOS quality class",
            "Maps delay and late-drop impairment to an objective voice-quality estimate.",
            "Summarize in the QoE table; keep the simplified E-model assumption explicit.",
        ),
        (
            "Bandwidth",
            "Average and peak relay payload bitrate",
            "Quantifies the cost of uncompressed PCM16 relay traffic without cluttering the workbook with raw byte counters.",
            "Use when discussing scalability as client count increases.",
        ),
        (
            "Relay stability",
            "Relay send drop rate",
            "Shows whether bounded per-recipient queues are dropping packets due to slow receivers.",
            "Use to explain server-side behavior in congested cases.",
        ),
        (
            "Browser runtime",
            "AudioWorklet callback stddev, resampled frames, capture queue drops",
            "Confirms capture timing stability and whether hardware sample-rate mismatch affected the test.",
            "Use as validity checks, not as the headline result.",
        ),
    ]
    for row_index, row in enumerate(rows, start=5):
        _write_row(sheet, row_index, row)


def _write_table(sheet: object, fields: list[str], rows: list[Mapping[str, object]], table_cls: object, style_cls: object, font_cls: object, fill_cls: object) -> None:
    _write_row(sheet, 4, fields)
    _style_header(sheet, f"A4:{_column_letter(len(fields))}4", font_cls, fill_cls)
    for row_index, row in enumerate(rows, start=5):
        _write_row(sheet, row_index, [row.get(field, 0) for field in fields])
    if rows:
        table_ref = f"A4:{_column_letter(len(fields))}{len(rows) + 4}"
        table_name = "".join(ch for ch in sheet.title if ch.isalnum())[:20] + "Table"
        table = table_cls(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = style_cls(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)


def _client_summary_row(client: ClientMetricState) -> MetricRow:
    latest = client.latest
    uptime = max(time.monotonic() - client.joined_at, 0.0)
    return {
        "client_id": client.client_id,
        "name": client.name,
        "room": client.room,
        "status": "active" if client.active else "disconnected",
        "uptime_seconds": round(uptime, 3),
        "session_duration_seconds": _as_float(latest.get("session_duration_seconds")),
        "captured_frames": _as_int(latest.get("captured_frames")),
        "sent_packets": _as_int(latest.get("sent_packets")),
        "sent_bytes": _as_int(latest.get("sent_bytes")),
        "sent_payload_bytes": _as_int(latest.get("sent_payload_bytes")),
        "received_packets": _as_int(latest.get("received_packets")),
        "received_bytes": _as_int(latest.get("received_bytes")),
        "received_payload_bytes": _as_int(latest.get("received_payload_bytes")),
        "played_packets": _as_int(latest.get("played_packets")),
        "capture_errors": _as_int(latest.get("capture_errors")),
        "malformed_audio_packets": _as_int(latest.get("malformed_audio_packets")),
        "network_loss_packets": _as_int(latest.get("network_loss_packets")),
        "late_dropped_packets": _as_int(latest.get("late_dropped_packets")),
        "queue_overflow_dropped_packets": _as_int(latest.get("queue_overflow_dropped_packets")),
        "buffer_underrun_events": _as_int(latest.get("buffer_underrun_events")),
        "buffer_underrun_seconds": _as_float(latest.get("buffer_underrun_seconds")),
        "max_buffer_underrun_ms": _as_float(latest.get("max_buffer_underrun_ms")),
        "rtt_ms": _as_float(latest.get("rtt_ms")),
        "estimated_owd_ms": _as_float(latest.get("estimated_owd_ms")),
        "rfc3550_jitter_ms": _as_float(latest.get("rfc3550_jitter_ms")),
        "estimated_playout_latency_ms": _as_float(latest.get("estimated_playout_latency_ms")),
        "late_drop_rate_percent": _as_float(latest.get("late_drop_rate_percent")),
        "buffer_underrun_rate_per_min": _as_float(latest.get("buffer_underrun_rate_per_min")),
        "r_factor": _as_float(latest.get("r_factor")),
        "estimated_mos": _as_float(latest.get("estimated_mos")),
        "mos_quality": str(latest.get("mos_quality") or ""),
        "audio_context_sample_rate": _as_float(latest.get("audio_context_sample_rate")),
        "resampled_frames": _as_int(latest.get("resampled_frames")),
        "capture_queue_dropped_frames": _as_int(latest.get("capture_queue_dropped_frames")),
        "active_remote_streams": _as_int(latest.get("active_remote_streams")),
        "callback_interval_stddev_ms": _as_float(latest.get("callback_interval_stddev_ms")),
        "worklet_message_interval_stddev_ms": _as_float(latest.get("worklet_message_interval_stddev_ms")),
        "last_sent_kbps": _as_float(latest.get("last_sent_kbps")),
        "last_rx_kbps": _as_float(latest.get("last_rx_kbps")),
        "playback_queue_seconds": _as_float(latest.get("playback_queue_seconds")),
    }


def _add_qos_chart(qos_summary: object, rows: list[MetricRow], chart_cls: object, reference_cls: object) -> None:
    if len(rows) < 2:
        return
    helper_start_row = 4
    helper_start_col = 7
    headers = ["Sample", "OWD ms", "Jitter ms", "Latency ms"]
    for offset, value in enumerate(headers):
        qos_summary.cell(helper_start_row, helper_start_col + offset, value)
    for row_index, sample in enumerate(rows, start=helper_start_row + 1):
        qos_summary.cell(row_index, helper_start_col, row_index - helper_start_row)
        qos_summary.cell(row_index, helper_start_col + 1, _as_float(sample.get("browser_avg_estimated_owd_ms")))
        qos_summary.cell(row_index, helper_start_col + 2, _as_float(sample.get("browser_max_jitter_ms")))
        qos_summary.cell(row_index, helper_start_col + 3, _as_float(sample.get("browser_avg_playout_latency_ms")))
    for column_index in range(helper_start_col, helper_start_col + len(headers)):
        qos_summary.column_dimensions[_column_letter(column_index)].hidden = True

    chart = chart_cls()
    _format_line_chart(chart, "QoS Delay Trend", "Milliseconds", "Sample")
    data = reference_cls(
        qos_summary,
        min_col=helper_start_col + 1,
        max_col=helper_start_col + 3,
        min_row=helper_start_row,
        max_row=helper_start_row + len(rows),
    )
    categories = reference_cls(
        qos_summary,
        min_col=helper_start_col,
        min_row=helper_start_row + 1,
        max_row=helper_start_row + len(rows),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    _style_chart_series(chart, ["4472C4", "70AD47", "C00000"])
    qos_summary.add_chart(chart, "F5")


def _format_line_chart(chart: object, title: str, y_title: str, x_title: str) -> None:
    chart.title = title
    chart.style = 2
    chart.height = 6.0
    chart.width = 11.5
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.legend.position = "r"
    chart.y_axis.numFmt = "0.0"
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None


def _style_chart_series(chart: object, colors: list[str]) -> None:
    for series, color in zip(chart.series, colors):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 22000


def _title(sheet: object, title: str, font_cls: object, fill_cls: object) -> None:
    from openpyxl.styles import Alignment

    sheet["A1"] = title
    sheet["A2"] = "Processed measurement workbook for browser-based secure LAN intercom."
    sheet["A1"].font = font_cls(size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = fill_cls("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"].font = font_cls(size=10, italic=True, color="404040")
    sheet["A2"].fill = fill_cls("solid", fgColor="EAF2F8")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.merge_cells("A1:H1")
    sheet.merge_cells("A2:H2")
    sheet.row_dimensions[1].height = 36
    sheet.row_dimensions[2].height = 34


def _style_header(sheet: object, range_name: str, font_cls: object, fill_cls: object) -> None:
    from openpyxl.styles import Alignment

    for row in sheet[range_name]:
        for cell in row:
            cell.font = font_cls(bold=True, color="FFFFFF")
            cell.fill = fill_cls("solid", fgColor="2F5597")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[cell.row].height = 22


def _size_columns(sheet: object, alignment_cls: object) -> None:
    from openpyxl.utils import get_column_letter

    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        if sheet.column_dimensions[letter].hidden:
            continue
        width = max(
            len(str(sheet.cell(row_index, column_index).value))
            if sheet.cell(row_index, column_index).value is not None
            else 0
            for row_index in range(1, sheet.max_row + 1)
        )
        cap = 72 if column_index in (4, 5) else 34
        sheet.column_dimensions[letter].width = min(max(width + 2, 12), cap)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row not in (1, 2, 4):
                cell.alignment = alignment_cls(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    if sheet.title in {"Summary", "QoS Summary", "QoE Summary", "Paper Metrics"}:
        for row_index in range(5, sheet.max_row + 1):
            longest = max(
                len(str(sheet.cell(row_index, column_index).value))
                if sheet.cell(row_index, column_index).value is not None
                else 0
                for column_index in range(1, sheet.max_column + 1)
            )
            if longest > 130:
                sheet.row_dimensions[row_index].height = 58
            elif longest > 80:
                sheet.row_dimensions[row_index].height = 42
            elif longest > 42:
                sheet.row_dimensions[row_index].height = 30
            else:
                sheet.row_dimensions[row_index].height = 21
    sheet.row_dimensions[1].height = 36
    sheet.row_dimensions[2].height = 34
    sheet.row_dimensions[3].height = 8


def _write_row(sheet: object, row_index: int, values: list[object] | tuple[object, ...]) -> None:
    for column_index, value in enumerate(values, start=1):
        sheet.cell(row_index, column_index, value)


def _clean_metric_value(value: object) -> MetricValue:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float, str)):
        return value
    return 0


def _as_float(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _as_int(value: object) -> int:
    return int(_as_float(value))


def _mb(value: object) -> float:
    return round(_as_float(value) / 1_000_000, 3)


def _append_positive(values: list[float], value: object) -> None:
    number = _as_float(value)
    if number > 0:
        values.append(number)


def _mean(values: list[float]) -> float:
    if not values:
        return 0.0
    return round(sum(values) / len(values), 3)


def _max_values(values: list[float]) -> float:
    if not values:
        return 0.0
    return round(max(values), 3)


def _min_values(values: list[float]) -> float:
    if not values:
        return 0.0
    return round(min(values), 3)


def _percentile(values: list[float], percentile: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return round(ordered[0], 3)
    rank = (len(ordered) - 1) * max(0.0, min(100.0, percentile)) / 100
    lower = int(rank)
    upper = min(lower + 1, len(ordered) - 1)
    fraction = rank - lower
    return round(ordered[lower] + (ordered[upper] - ordered[lower]) * fraction, 3)


def _client_sample_values(clients: list[ClientMetricState], field: str) -> list[float]:
    values: list[float] = []
    for client in clients:
        for sample in client.samples:
            value = _as_float(sample.get(field))
            if value > 0:
                values.append(value)
    return values


def _max(rows: list[MetricRow], field: str) -> float:
    if not rows:
        return 0.0
    return round(max(_as_float(row.get(field)) for row in rows), 3)


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"
